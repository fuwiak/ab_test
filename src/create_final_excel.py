"""
Создание итогового Excel файла со всеми сводными таблицами
для финальной работы по курсу
"""

import pandas as pd
import numpy as np
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def create_comprehensive_excel_report():
    """Создание полного Excel отчёта со всеми таблицами"""
    
    # Создание Excel writer
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"../reports/Финальная_работа_AB_тест_{timestamp}.xlsx"
    
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        
        # Лист 1: Основные результаты
        create_main_results_sheet(writer)
        
        # Лист 2: ARPU по группам и платформам
        create_arpu_analysis_sheet(writer)
        
        # Лист 3: Доверительные интервалы
        create_confidence_intervals_sheet(writer)
        
        # Лист 4: Статистические тесты
        create_statistical_tests_sheet(writer)
        
        # Лист 5: Очистка данных
        create_data_cleaning_sheet(writer)
        
        # Лист 6: Проекция дохода
        create_revenue_projection_sheet(writer)
        
        # Лист 7: Бизнес-рекомендации
        create_business_recommendations_sheet(writer)
    
    # Форматирование Excel файла
    format_excel_file(filename)
    
    print(f"✅ Создан итоговый Excel файл: {filename}")
    return filename

def create_main_results_sheet(writer):
    """Лист 1: Основные результаты анализа"""
    
    # Основная сводная таблица
    main_results = pd.DataFrame({
        'Метрика': ['ARPU (USD)', 'ARPPU (USD)', 'Траты валюты (монеты)'],
        'Контрольная группа': [5.8295, 5.8311, 5800.71],
        'Тестовая группа': [6.1622, 6.1631, 6229.57],
        'Абсолютное изменение': [0.3327, 0.3320, 428.86],
        'Относительное изменение (%)': [5.71, 5.69, 7.39],
        'p-значение': ['< 0.000001', '< 0.000001', '< 0.000001'],
        'Статистически значимо': ['✅ ДА', '✅ ДА', '✅ ДА'],
        'Практически значимо': ['✅ ДА', '✅ ДА', '✅ ДА']
    })
    
    main_results.to_excel(writer, sheet_name='Основные результаты', index=False)
    
    # Информация о выборке
    sample_info = pd.DataFrame({
        'Показатель': [
            'Исходное количество игроков',
            'Удалено читеров',
            'Удалено статистических выбросов',
            'Финальная выборка',
            'Контрольная группа',
            'Тестовая группа',
            'Процент сохранённых данных'
        ],
        'Значение': [
            '8,640,000',
            '353 (0.004%)',
            '346 (0.004%)',
            '8,634,408',
            '4,319,928 (50.0%)',
            '4,314,480 (50.0%)',
            '99.935%'
        ]
    })
    
    # Добавление информации о выборке на тот же лист
    startrow = len(main_results) + 3
    sample_info.to_excel(writer, sheet_name='Основные результаты', 
                        startrow=startrow, index=False)

def create_arpu_analysis_sheet(writer):
    """Лист 2: ARPU по группам и платформам"""
    
    # ARPU по платформам и группам
    platform_arpu = pd.DataFrame({
        'Платформа': ['PC', 'PC', 'PS4', 'PS4', 'Xbox', 'Xbox'],
        'Группа': ['Control', 'Test', 'Control', 'Test', 'Control', 'Test'],
        'Количество игроков': [1150285, 1150842, 1150746, 1148250, 1154912, 1152493],
        'ARPU (USD)': [5.6462, 6.2690, 5.7376, 6.0848, 6.1035, 6.1328],
        'Стандартное отклонение': [1.8168, 1.9295, 1.8671, 1.8849, 1.9044, 1.9128],
        'Медиана': [5.95, 5.95, 5.95, 5.95, 5.95, 5.95]
    })
    
    platform_arpu.to_excel(writer, sheet_name='ARPU по платформам', index=False)
    
    # Улучшения по платформам
    platform_improvements = pd.DataFrame({
        'Платформа': ['PC', 'PS4', 'Xbox'],
        'ARPU Control': [5.6462, 5.7376, 6.1035],
        'ARPU Test': [6.2690, 6.0848, 6.1328],
        'Абсолютное улучшение': [0.6228, 0.3472, 0.0293],
        'Относительное улучшение (%)': [11.02, 6.05, 0.48],
        'Эффективность': ['Высокая', 'Средняя', 'Низкая']
    })
    
    startrow = len(platform_arpu) + 3
    platform_improvements.to_excel(writer, sheet_name='ARPU по платформам', 
                                  startrow=startrow, index=False)

def create_confidence_intervals_sheet(writer):
    """Лист 3: 95% Доверительные интервалы"""
    
    # Доверительные интервалы для ARPU
    arpu_ci = pd.DataFrame({
        'Метрика': ['ARPU', 'ARPU'],
        'Группа': ['Control', 'Test'],
        'Среднее значение': [5.8295, 6.1622],
        'Стандартная ошибка': [0.0006, 0.0006],
        'ДИ нижняя граница': [5.8289, 6.1616],
        'ДИ верхняя граница': [5.8301, 6.1628],
        'Погрешность (±)': [0.0006, 0.0006],
        'Размер выборки': [4319928, 4314480]
    })
    
    arpu_ci.to_excel(writer, sheet_name='Доверительные интервалы', index=False)
    
    # Доверительные интервалы для ARPPU
    arppu_ci = pd.DataFrame({
        'Метрика': ['ARPPU', 'ARPPU'],
        'Группа': ['Control', 'Test'],
        'Среднее значение': [5.8311, 6.1631],
        'Стандартная ошибка': [0.0006, 0.0006],
        'ДИ нижняя граница': [5.8305, 6.1625],
        'ДИ верхняя граница': [5.8317, 6.1637],
        'Погрешность (±)': [0.0006, 0.0006],
        'Размер выборки': [4318720, 4313872]
    })
    
    startrow = len(arpu_ci) + 2
    arppu_ci.to_excel(writer, sheet_name='Доверительные интервалы', 
                     startrow=startrow, index=False)
    
    # Доверительные интервалы для трат валюты
    cash_ci = pd.DataFrame({
        'Метрика': ['Траты валюты', 'Траты валюты'],
        'Группа': ['Control', 'Test'],
        'Среднее значение': [5800.71, 6229.57],
        'Стандартная ошибка': [1.27, 1.34],
        'ДИ нижняя граница': [5799.44, 6228.23],
        'ДИ верхняя граница': [5801.98, 6230.91],
        'Погрешность (±)': [1.27, 1.34],
        'Размер выборки': [4319928, 4314480]
    })
    
    startrow = len(arpu_ci) + len(arppu_ci) + 4
    cash_ci.to_excel(writer, sheet_name='Доверительные интервалы', 
                    startrow=startrow, index=False)

def create_statistical_tests_sheet(writer):
    """Лист 4: Статистические тесты"""
    
    statistical_results = pd.DataFrame({
        'Метрика': ['ARPU', 'ARPPU', 'Траты валюты'],
        'Control среднее': [5.8295, 5.8311, 5800.71],
        'Test среднее': [6.1622, 6.1631, 6229.57],
        'Улучшение (%)': [5.71, 5.69, 7.39],
        't-статистика': [-258.37, -257.99, -456.73],
        'p-значение': ['< 0.000001', '< 0.000001', '< 0.000001'],
        "Cohen's d": [-0.179, -0.179, -0.312],
        'Размер эффекта': ['Малый', 'Малый', 'Малый'],
        'Статистически значимо': ['✅ ДА', '✅ ДА', '✅ ДА'],
        'Высоко значимо (p<0.001)': ['✅ ДА', '✅ ДА', '✅ ДА']
    })
    
    statistical_results.to_excel(writer, sheet_name='Статистические тесты', index=False)
    
    # Интерпретация результатов
    interpretation = pd.DataFrame({
        'Критерий оценки': [
            'Статистическая значимость',
            'Практическая значимость',
            'Пересечение доверительных интервалов',
            'Размер выборки',
            'Качество данных',
            'Консистентность результатов'
        ],
        'Результат': [
            '✅ Все метрики p < 0.000001',
            '✅ Улучшения 5-7% практически значимы',
            '❌ Интервалы НЕ пересекаются',
            '✅ 8.6М игроков - отличная мощность',
            '✅ 99.9% данных после очистки',
            '✅ Положительный эффект на всех платформах'
        ],
        'Оценка': [
            'ОТЛИЧНО',
            'ОТЛИЧНО', 
            'ОТЛИЧНО',
            'ОТЛИЧНО',
            'ОТЛИЧНО',
            'ОТЛИЧНО'
        ]
    })
    
    startrow = len(statistical_results) + 3
    interpretation.to_excel(writer, sheet_name='Статистические тесты', 
                           startrow=startrow, index=False)

def create_data_cleaning_sheet(writer):
    """Лист 5: Процесс очистки данных"""
    
    # Этапы очистки
    cleaning_steps = pd.DataFrame({
        'Этап': [
            'Исходные данные',
            'После удаления известных читеров',
            'После удаления статистических выбросов',
            'Финальная выборка'
        ],
        'Количество игроков': [8640000, 8637176, 8634408, 8634408],
        'Удалено на этапе': [0, 2824, 2768, 0],
        'Кумулятивно удалено': [0, 2824, 5592, 5592],
        'Процент сохранено': [100.0, 99.967, 99.935, 99.935]
    })
    
    cleaning_steps.to_excel(writer, sheet_name='Очистка данных', index=False)
    
    # Методы выявления аномалий
    methods = pd.DataFrame({
        'Тип аномалий': ['Известные читеры', 'Статистические выбросы'],
        'Метод выявления': ['Флаг cheaters = 1', 'IQR метод (Q3 + 3×IQR)'],
        'Критерий': ['Прямое указание в данных', 'Траты > 12,650 монет'],
        'Выявлено': [353, 346],
        'Процент от выборки': ['0.004%', '0.004%'],
        'Обоснование удаления': [
            'Искажают реальные метрики',
            'Аномальные траты (в 2+ раза выше нормы)'
        ]
    })
    
    startrow = len(cleaning_steps) + 3
    methods.to_excel(writer, sheet_name='Очистка данных', 
                    startrow=startrow, index=False)

def create_revenue_projection_sheet(writer):
    """Лист 6: Проекция увеличения дохода"""
    
    # Проекция дохода
    revenue_projection = pd.DataFrame({
        'Период': ['Дневной', 'Недельный', 'Месячный', 'Квартальный', 'Годовой'],
        'Базовый доход (млн USD)': [50.3, 352.1, 1509, 4527, 18359],
        'Увеличение дохода (млн USD)': [2.87, 20.09, 86.1, 258.5, 1048.7],
        'Новый доход (млн USD)': [53.17, 372.19, 1595.1, 4785.5, 19407.7],
        'Прирост (%)': [5.71, 5.71, 5.71, 5.71, 5.71]
    })
    
    revenue_projection.to_excel(writer, sheet_name='Проекция дохода', index=False)
    
    # Допущения и расчёты
    assumptions = pd.DataFrame({
        'Параметр': [
            'Базовый ARPU',
            'Улучшение ARPU',
            'Количество игроков',
            'Дневной базовый доход',
            'Дневное улучшение',
            'Валидность проекции'
        ],
        'Значение': [
            '$5.8295',
            '+5.71% (+$0.3327)',
            '8,634,408',
            '$50.3 млн',
            '+$2.87 млн',
            'На основе статистически значимых данных'
        ],
        'Источник': [
            'Анализ контрольной группы',
            'Статистический тест (p<0.000001)',
            'Финальная выборка после очистки',
            'ARPU × Количество игроков',
            'Улучшение × Количество игроков',
            'Доверительный интервал 95%'
        ]
    })
    
    startrow = len(revenue_projection) + 3
    assumptions.to_excel(writer, sheet_name='Проекция дохода', 
                        startrow=startrow, index=False)

def create_business_recommendations_sheet(writer):
    """Лист 7: Бизнес-рекомендации"""
    
    # Основная рекомендация
    main_recommendation = pd.DataFrame({
        'Аспект': ['ОСНОВНАЯ РЕКОМЕНДАЦИЯ'],
        'Решение': ['✅ ВНЕДРИТЬ АКЦИЮ НА ПОСТОЯННОЙ ОСНОВЕ'],
        'Обоснование': ['Все ключевые метрики показывают статистически и практически значимые улучшения'],
        'Ожидаемый эффект': ['+5.7% увеличение дохода, +7.4% активности игроков']
    })
    
    main_recommendation.to_excel(writer, sheet_name='Бизнес-рекомендации', index=False)
    
    # Детальные рекомендации
    detailed_recommendations = pd.DataFrame({
        'Направление': [
            'Внедрение',
            'Мониторинг',
            'Оптимизация',
            'Расширение',
            'Риски'
        ],
        'Рекомендация': [
            'Поэтапное внедрение: начать с PC (наибольший эффект +11%)',
            'Отслеживать KPI первые 3 месяца после запуска',
            'Тестировать размер скидки и частоту акций',
            'Применить подход к другим игровым элементам',
            'Низкий риск благодаря высокой статистической значимости'
        ],
        'Временные рамки': [
            '1-2 месяца',
            '3 месяца',
            '6 месяцев',
            '12 месяцев',
            'Постоянно'
        ],
        'Ответственность': [
            'Product Manager',
            'Data Analytics',
            'Game Design',
            'Product Strategy',
            'Risk Management'
        ]
    })
    
    startrow = len(main_recommendation) + 3
    detailed_recommendations.to_excel(writer, sheet_name='Бизнес-рекомендации', 
                                    startrow=startrow, index=False)
    
    # KPI для мониторинга
    kpi_monitoring = pd.DataFrame({
        'KPI': [
            'ARPU',
            'ARPPU', 
            'Конверсия в покупки',
            'Retention Rate',
            'Траты внутриигровой валюты',
            'NPS (Net Promoter Score)'
        ],
        'Текущее значение': [
            '$6.16 (тест группа)',
            '$6.16 (тест группа)',
            'Требует измерения',
            'Требует измерения',
            '6,230 монет (тест группа)',
            'Требует измерения'
        ],
        'Целевое изменение': [
            'Сохранить +5.7%',
            'Сохранить +5.7%',
            'Не снижать',
            'Не снижать',
            'Сохранить +7.4%',
            'Улучшить'
        ],
        'Частота измерения': [
            'Еженедельно',
            'Еженедельно',
            'Еженедельно',
            'Ежемесячно',
            'Еженедельно',
            'Ежемесячно'
        ]
    })
    
    startrow = len(main_recommendation) + len(detailed_recommendations) + 6
    kpi_monitoring.to_excel(writer, sheet_name='Бизнес-рекомендации', 
                           startrow=startrow, index=False)

def format_excel_file(filename):
    """Форматирование Excel файла для лучшей читаемости"""
    
    # Загрузка workbook
    wb = openpyxl.load_workbook(filename)
    
    # Стили
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    alignment_center = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'), 
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Применение стилей ко всем листам
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Автоширина колонок
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Форматирование заголовков
        for row in ws.iter_rows(min_row=1, max_row=1):
            for cell in row:
                if cell.value:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = alignment_center
                    cell.border = thin_border
        
        # Границы для всех ячеек с данными
        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    cell.border = thin_border
    
    # Сохранение отформатированного файла
    wb.save(filename)

def main():
    """Главная функция создания Excel отчёта"""
    
    print("📊 Создание итогового Excel файла для финальной работы...")
    print("=" * 70)
    
    filename = create_comprehensive_excel_report()
    
    print("\n📋 Созданные листы:")
    print("   1. Основные результаты - главные метрики и выборка")
    print("   2. ARPU по платформам - детальный анализ по PC/PS4/Xbox")
    print("   3. Доверительные интервалы - 95% ДИ для всех метрик")
    print("   4. Статистические тесты - t-тесты и размеры эффекта")
    print("   5. Очистка данных - методы и результаты очистки")
    print("   6. Проекция дохода - финансовые прогнозы")
    print("   7. Бизнес-рекомендации - выводы и план действий")
    
    print(f"\n✅ Excel файл готов: {filename}")
    print("📎 Готов к включению в финальную работу по курсу!")

if __name__ == "__main__":
    main() 